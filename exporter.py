# -*- coding: utf-8 -*-

# import sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook


def save2xlsx(data):
    print()
    print('...............................................................')
    print(f'Начинаю писать в report_{datetime.date(datetime.today())}.xlsx')
    wb = Workbook()
    wb['Sheet'].title = 'Tinkoff Report'
    sheet = wb.active
    # Название отчета
    sheet['A1'] = 'Отчет на основе анализа данных Tinkoff Инвестиции'
    # Шапка таблицы
    sheet['B2'] = 'Название'
    sheet['C2'] = 'Тикер'
    sheet['D2'] = 'Кол-во входов'
    sheet['E2'] = 'Купленно всего'
    sheet['F2'] = 'Сейчас в портфеле'
    sheet['G2'] = 'Вложено'
    sheet['H2'] = 'Получено'
    sheet['I2'] = 'Доход'

    sheet['J2'] = 'Время владения'
    sheet['K2'] = 'Прибыль на акцию'
    sheet['L2'] = 'Эффективность %'
    sheet['M2'] = 'gIndex'
    sheet['N2'] = ''
    sheet['O2'] = 'Сейчас вложено'
    sheet['P2'] = 'Средняя цена'

    x = 3
    for ticker in data:
        delta = 1 if data[ticker]['Time']['Delta'] is None else data[ticker]['Time']['Delta']
        print(ticker, ' - обработан и записан в XLSX')
        sheet.cell(row=x, column=2).value = data[ticker]['Name']
        sheet.cell(row=x, column=3).value = ticker
        sheet.cell(row=x, column=4).value = data[ticker]['OutPosition']
        sheet.cell(row=x, column=5).value = data[ticker]['Qty']['PositiveSummary']
        sheet.cell(row=x, column=6).value = data[ticker]['Qty']['Summary']
        if data[ticker]['OnHands']['Amount'] > 0:
            sheet.cell(row=x, column=7).value = data[ticker]['Amount']['SummaryBay']
            sheet.cell(row=x, column=8).value = data[ticker]['Amount']['SummarySell'] - data[ticker]['OnHands']['Amount']
            sheet.cell(row=x, column=9).value = data[ticker]['Amount']['Summary'] - data[ticker]['OnHands']['Amount']
        elif data[ticker]['OnHands']['Amount'] < 0:
            sheet.cell(row=x, column=7).value = data[ticker]['Amount']['SummaryBay'] - data[ticker]['OnHands']['Amount']
            sheet.cell(row=x, column=8).value = data[ticker]['Amount']['SummarySell']
            sheet.cell(row=x, column=9).value = data[ticker]['Amount']['Summary'] - data[ticker]['OnHands']['Amount']
        elif data[ticker]['OnHands']['Amount'] == 0:
            sheet.cell(row=x, column=7).value = data[ticker]['Amount']['SummaryBay']
            sheet.cell(row=x, column=8).value = data[ticker]['Amount']['SummarySell']
            sheet.cell(row=x, column=9).value = data[ticker]['Amount']['Summary']

        sheet.cell(row=x, column=10).value = delta
        sheet.cell(row=x, column=11).value = round(data[ticker]['Amount']['Summary'] / data[ticker]['Qty']['PositiveSummary'], 2)
        sheet.cell(row=x, column=12).value = 'в разработке'
        sheet.cell(row=x, column=13).value = 'в разработке'

        sheet.cell(row=x, column=15).value = data[ticker]['OnHands']['Amount']
        if data[ticker]['Qty']['Summary'] != 0:
            sheet.cell(row=x, column=15).value = data[ticker]['OnHands']['Amount']
            sheet.cell(row=x, column=16).value = round(data[ticker]['OnHands']['Amount'] * -1 / data[ticker]['Qty']['Summary'], 2)
        else:
            sheet.cell(row=x, column=15).value = data[ticker]['OnHands']['Amount'] * -1
            sheet.cell(row=x, column=16).value = 0
        x += 1
    wb.save(f'report_{datetime.date(datetime.today())}.xlsx')
    wb.close()
    print()
    print('......................................................')
    print('Все готово!')
